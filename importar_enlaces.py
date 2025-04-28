import openpyxl
import sqlite3

def importar_enlaces(ruta_archivo, db_path):
    """
    Importar enlaces desde un archivo Excel y guardarlos en la base de datos.
    """
    try:
        # Abrir el archivo Excel
        workbook = openpyxl.load_workbook(ruta_archivo)
        sheet = workbook.active  # Usar la hoja activa

        # Establecer conexión a la base de datos
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Leer e insertar enlaces
        enlaces = []
        pregunta = sheet.cell(row=1, column=1).value
        for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezados
            enlace = row[1]  # Asumimos que el enlace está en la primera columna
            titulo =row[1]
            if enlace:
                enlaces.append(enlace)
                cursor.execute("INSERT INTO documentos (enlace, title,cumplimiento_de_criterios) VALUES (?,?,?)", (enlace,pregunta,pregunta))

        conn.commit()
        conn.close()

        print(f"Importados {len(enlaces)} enlaces desde {ruta_archivo}.")
        return len(enlaces)

    except Exception as e:
        print(f"Error al importar enlaces: {str(e)}")
        return 0